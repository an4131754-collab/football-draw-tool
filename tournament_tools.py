from __future__ import annotations

import io
import json
import math
import secrets
import shutil
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from itertools import combinations
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_ROOT = BASE_DIR / "outputs"
ARCHIVE_ROOT = OUTPUT_ROOT / "archive"
LATEST_ROOT = OUTPUT_ROOT / "latest"

CONFIG_FILENAME = "config.json"
DRAW_RESULT_FILENAME = "draw_result.json"
SCHEDULE_FILENAME = "114學年度足球系際盃_賽程表.xlsx"
PDF_FILENAME = "114學年度足球系際盃_抽籤隨機性說明.pdf"
DRAW_TITLE = "114學年度足球系際盃賽程表"

RANDOM_FUNCTION_NAME = "secrets.SystemRandom().shuffle"
RANDOM_FUNCTION_SUMMARY = "使用作業系統提供的安全亂數來源進行洗牌"

GROUP_LABELS = tuple(chr(ord("A") + index) for index in range(26))
TEMPLATE_GROUPS = ("A", "B", "C", "D")
AVAILABLE_OUTPUTS = ("json", "schedule", "pdf")
DEFAULT_DOWNLOAD_OPTIONS = {"json": True, "schedule": True, "pdf": True}
REFEREE_SHEET_NAME = "裁判"

GROUP_SUMMARY_CELLS = {
    "A": ("L3", "M3", "N3", "O3"),
    "B": ("L5", "M5", "N5", "O5"),
    "C": ("L7", "M7", "N7", "O7"),
    "D": ("L9", "M9", "N9", "O9"),
}

GROUP_MATCH_CELLS = {
    "A": (("D5", "E5"), ("D7", "E7"), ("D9", "E9")),
    "B": (("I4", "J4"), ("I6", "J6"), ("I8", "J8")),
    "C": (("D4", "E4"), ("D6", "E6"), ("D8", "E8")),
    "D": (("I5", "J5"), ("I7", "J7"), ("I9", "J9")),
}

GROUP_MATCH_ORDERS = {
    "A": ((1, 2), (2, 0), (0, 1)),
    "B": ((0, 1), (1, 2), (2, 0)),
    "C": ((0, 1), (1, 2), (2, 0)),
    "D": ((0, 1), (1, 2), (2, 0)),
}

TEMPLATE_SCHEDULE_MATCH_CELLS = (
    {"day": "DAY1", "time_cell": "A4", "match_cell": "B4", "field": "甲", "group_cell": "C4", "home_cell": "D4", "away_cell": "E4"},
    {"day": "DAY1", "time_cell": "A4", "match_cell": "G4", "field": "乙", "group_cell": "H4", "home_cell": "I4", "away_cell": "J4"},
    {"day": "DAY1", "time_cell": "A5", "match_cell": "B5", "field": "甲", "group_cell": "C5", "home_cell": "D5", "away_cell": "E5"},
    {"day": "DAY1", "time_cell": "A5", "match_cell": "G5", "field": "乙", "group_cell": "H5", "home_cell": "I5", "away_cell": "J5"},
    {"day": "DAY1", "time_cell": "A6", "match_cell": "B6", "field": "甲", "group_cell": "C6", "home_cell": "D6", "away_cell": "E6"},
    {"day": "DAY1", "time_cell": "A6", "match_cell": "G6", "field": "乙", "group_cell": "H6", "home_cell": "I6", "away_cell": "J6"},
    {"day": "DAY1", "time_cell": "A7", "match_cell": "B7", "field": "甲", "group_cell": "C7", "home_cell": "D7", "away_cell": "E7"},
    {"day": "DAY1", "time_cell": "A7", "match_cell": "G7", "field": "乙", "group_cell": "H7", "home_cell": "I7", "away_cell": "J7"},
    {"day": "DAY1", "time_cell": "A8", "match_cell": "B8", "field": "甲", "group_cell": "C8", "home_cell": "D8", "away_cell": "E8"},
    {"day": "DAY1", "time_cell": "A8", "match_cell": "G8", "field": "乙", "group_cell": "H8", "home_cell": "I8", "away_cell": "J8"},
    {"day": "DAY1", "time_cell": "A9", "match_cell": "B9", "field": "甲", "group_cell": "C9", "home_cell": "D9", "away_cell": "E9"},
    {"day": "DAY1", "time_cell": "A9", "match_cell": "G9", "field": "乙", "group_cell": "H9", "home_cell": "I9", "away_cell": "J9"},
    {"day": "DAY2", "time_cell": "A14", "match_cell": "B14", "field": "甲", "group_cell": "", "home_cell": "D14", "away_cell": "E14"},
    {"day": "DAY2", "time_cell": "A14", "match_cell": "G14", "field": "乙", "group_cell": "", "home_cell": "I14", "away_cell": "J14"},
    {"day": "DAY2", "time_cell": "A15", "match_cell": "B15", "field": "甲", "group_cell": "", "home_cell": "D15", "away_cell": "E15"},
    {"day": "DAY2", "time_cell": "A15", "match_cell": "G15", "field": "乙", "group_cell": "", "home_cell": "I15", "away_cell": "J15"},
    {"day": "DAY2", "time_cell": "A17", "match_cell": "B17", "field": "甲", "group_cell": "", "home_cell": "D17", "away_cell": "E17"},
    {"day": "DAY2", "time_cell": "A17", "match_cell": "G17", "field": "乙", "group_cell": "", "home_cell": "I17", "away_cell": "J17"},
    {"day": "DAY2", "time_cell": "A18", "match_cell": "B18", "field": "甲", "group_cell": "", "home_cell": "D18", "away_cell": "E18"},
    {"day": "DAY2", "time_cell": "A18", "match_cell": "G18", "field": "乙", "group_cell": "", "home_cell": "I18", "away_cell": "J18"},
)

PREFERRED_REGISTRATION_NAMES = (
    "114學年度足球⚽️系際盃 (回覆).xlsx",
    "114學年度足球系際盃 (回覆).xlsx",
)
PREFERRED_TEMPLATE_NAMES = ("113系際盃賽程及裁判表_公開版.xlsx",)

PDF_FONT_NAME = "TournamentTC"
PDF_FONT_CANDIDATES = (
    Path(r"C:\Windows\Fonts\NotoSansTC-VF.ttf"),
    Path(r"C:\Windows\Fonts\kaiu.ttf"),
)

DEFAULT_CONFIG: dict[str, Any] = {
    "registration_filenames": list(PREFERRED_REGISTRATION_NAMES),
    "template_filenames": list(PREFERRED_TEMPLATE_NAMES),
    "team_column": "科系",
    "min_team_count": 2,
    "default_group_count": 4,
    "max_group_count": 26,
    "default_advance_per_group": 1,
    "default_wildcard_count": 0,
    "default_knockout_format": "semifinal",
    "template_schedule_team_count": 12,
    "template_schedule_group_count": 4,
    "schedule_filename": SCHEDULE_FILENAME,
    "pdf_filename": PDF_FILENAME,
    "schedule_title": DRAW_TITLE,
    "group_label_format": "{group}組",
    "semifinal_labels": {
        "D14": "A1",
        "E14": "C1",
        "I14": "B1",
        "J14": "D1",
    },
    "final_labels": {
        "B17": 15,
        "G17": 16,
        "D17": "13L",
        "E17": "14L",
        "I17": "13W",
        "J17": "14W",
    },
    "clear_cells": ["B15", "D15", "E15", "G15", "I15", "J15"],
    "reserve_cells": ["B18", "D18", "E18", "G18", "I18", "J18"],
    "pdf_title": "114學年度足球系際盃抽籤隨機性說明",
    "fields": ["甲", "乙"],
    "referees_per_match": 3,
    "referee_sheet_name": REFEREE_SHEET_NAME,
    "match_duration_minutes": 45,
    "max_matches_per_team_per_day": 3,
    "final_preferred_start": "14:00",
    "default_day1_latest_end": "16:45",
    "default_day2_latest_end": "15:45",
    "day_slots": {
        "DAY1": ["09:00", "10:00", "11:00", "14:00", "15:00", "16:00"],
        "DAY2": ["10:00", "11:00", "14:00", "15:00"],
    },
    "latest_end_options": {
        "DAY1": ["16:45", "17:45", "18:45", "19:45", "20:45"],
        "DAY2": ["15:45", "16:45", "17:45", "18:45", "19:45", "20:45"],
    },
}


@dataclass(frozen=True)
class ArtifactPaths:
    output_dir: Path
    latest_dir: Path
    draw_json_path: Path
    schedule_path: Path | None
    pdf_path: Path | None
    latest_draw_json_path: Path
    latest_schedule_path: Path | None
    latest_pdf_path: Path | None
    latest_sync_complete: bool


def load_config(base_dir: Path = BASE_DIR) -> dict[str, Any]:
    config_path = base_dir / CONFIG_FILENAME
    config = deepcopy(DEFAULT_CONFIG)
    if not config_path.exists():
        return config

    with config_path.open("r", encoding="utf-8") as file_handle:
        user_config = json.load(file_handle)

    return merge_config(config, user_config)


def merge_config(default: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    result = deepcopy(default)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = merge_config(result[key], value)
        else:
            result[key] = value
    return result


def get_artifact_filenames(base_dir: Path = BASE_DIR) -> dict[str, str]:
    config = load_config(base_dir)
    return {
        "json": DRAW_RESULT_FILENAME,
        "schedule": str(config["schedule_filename"]),
        "pdf": str(config["pdf_filename"]),
    }


def resolve_registration_path(base_dir: Path = BASE_DIR) -> Path:
    config = load_config(base_dir)
    return _resolve_workbook(
        base_dir,
        preferred_names=tuple(config["registration_filenames"]),
        fallback_tokens=(".xlsx",),
        description="報名表",
        require_header=str(config["team_column"]),
    )


def resolve_template_path(base_dir: Path = BASE_DIR) -> Path:
    config = load_config(base_dir)
    return _resolve_workbook(
        base_dir,
        preferred_names=tuple(config["template_filenames"]),
        fallback_tokens=("113", "公開", ".xlsx"),
        description="公開版賽程模板",
    )


def _resolve_workbook(
    base_dir: Path,
    preferred_names: tuple[str, ...],
    fallback_tokens: tuple[str, ...],
    description: str,
    require_header: str | None = None,
) -> Path:
    for name in preferred_names:
        candidate = base_dir / name
        if candidate.exists() and (require_header is None or workbook_has_header(candidate, require_header)):
            return candidate

    candidates = sorted(base_dir.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    for candidate in candidates:
        name = candidate.name
        if all(token in name for token in fallback_tokens):
            if require_header is None or workbook_has_header(candidate, require_header):
                return candidate

    raise FileNotFoundError(f"找不到{description}，請確認檔案位於 {base_dir}")


def workbook_has_header(workbook_path: Path, header_name: str) -> bool:
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        header_row = next(workbook.worksheets[0].iter_rows(min_row=1, max_row=1, values_only=True))
    except Exception:
        return False
    return _find_header_index(header_row, header_name) is not None


def load_teams(
    registration_source: Path | str | BinaryIO,
    *,
    config: dict[str, Any] | None = None,
) -> list[str]:
    config = config or load_config(BASE_DIR)
    team_column_name = str(config["team_column"])
    workbook = load_workbook(registration_source, data_only=True)
    sheet = workbook.worksheets[0]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    team_column_index = _find_header_index(header_row, team_column_name)
    if team_column_index is None:
        raise ValueError(f"報名表找不到「{team_column_name}」欄位，無法讀取隊伍名單。")

    teams: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_value = row[team_column_index]
        if raw_value is None:
            continue

        team_name = normalize_team_name(raw_value)
        if team_name:
            teams.append(team_name)

    validate_teams(teams, config)
    return teams


def validate_teams(teams: list[str], config: dict[str, Any]) -> None:
    min_team_count = int(config["min_team_count"])
    if len(teams) < min_team_count:
        raise ValueError(f"報名表只有 {len(teams)} 隊，至少需要 {min_team_count} 隊才能抽籤。")

    duplicates = sorted({team for team in teams if teams.count(team) > 1})
    if duplicates:
        duplicate_list = "、".join(duplicates)
        raise ValueError(f"報名表有重複隊名：{duplicate_list}")


def _find_header_index(header_row: tuple[Any, ...], expected_header: str) -> int | None:
    for index, value in enumerate(header_row):
        if normalize_team_name(value) == expected_header:
            return index
    return None


def normalize_team_name(value: Any) -> str:
    return " ".join(str(value).strip().split())


def create_draw_data(
    teams: list[str],
    source_file: str,
    *,
    group_count: int,
    advance_per_group: int,
    wildcard_count: int,
    knockout_format: str | None = None,
    config: dict[str, Any] | None = None,
    download_options: dict[str, bool] | None = None,
    day1_latest_end: str | None = None,
    day2_latest_end: str | None = None,
) -> dict[str, Any]:
    config = config or load_config(BASE_DIR)
    knockout_format = normalize_knockout_format(knockout_format or str(config.get("default_knockout_format", "semifinal")))
    validate_draw_options(teams, group_count, advance_per_group, wildcard_count, knockout_format, config)

    shuffled_teams = list(teams)
    rng = secrets.SystemRandom()
    rng.shuffle(shuffled_teams)

    group_names = list(GROUP_LABELS[:group_count])
    groups: dict[str, list[str]] = {group_name: [] for group_name in group_names}
    slots: dict[str, str] = {}
    slot_order: list[str] = []

    for index, team in enumerate(shuffled_teams):
        group_name = group_names[index % group_count]
        groups[group_name].append(team)
        slot_name = f"{group_name}{len(groups[group_name])}"
        slots[slot_name] = team
        slot_order.append(slot_name)

    advancement = build_advancement_data(groups, advance_per_group, wildcard_count, knockout_format)

    draw_data = {
        "drawn_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": source_file,
        "team_count": len(teams),
        "group_count": group_count,
        "teams": teams,
        "slot_order": slot_order,
        "slots": slots,
        "groups": groups,
        "advancement": advancement,
        "knockout_format": knockout_format,
        "schedule_mode": determine_schedule_mode(len(teams), group_count, config),
        "download_options": normalize_download_options(download_options),
        "random_function": RANDOM_FUNCTION_NAME,
        "random_source": RANDOM_FUNCTION_SUMMARY,
    }
    draw_data["schedule"] = build_schedule_data(
        draw_data,
        config,
        day1_latest_end=day1_latest_end,
        day2_latest_end=day2_latest_end,
    )
    return draw_data


def validate_draw_options(
    teams: list[str],
    group_count: int,
    advance_per_group: int,
    wildcard_count: int,
    knockout_format: str,
    config: dict[str, Any],
) -> None:
    validate_teams(teams, config)

    max_group_count = int(config["max_group_count"])
    if group_count < 2:
        raise ValueError("組數至少需要 2 組。")
    if group_count > len(teams):
        raise ValueError("組數不能大於隊伍數。")
    if group_count > max_group_count:
        raise ValueError(f"目前最多支援 {max_group_count} 組。")

    if advance_per_group < 1:
        raise ValueError("每組晉級名額至少需要 1 名。")
    if wildcard_count < 0:
        raise ValueError("最佳名次補位數不能是負數。")

    smallest_group_size = len(teams) // group_count
    if advance_per_group > smallest_group_size:
        raise ValueError("每組晉級名額不能大於最小組別的隊伍數。")

    total_advancers = (group_count * advance_per_group) + wildcard_count
    if total_advancers > len(teams):
        raise ValueError("總晉級隊數不能大於參賽隊伍數。")
    required_advancers = knockout_advancer_count(knockout_format)
    if total_advancers != required_advancers:
        stage_label = knockout_format_label(knockout_format)
        raise ValueError(
            f"{stage_label} 需要剛好 {required_advancers} 隊晉級；"
            f"目前設定會產生 {total_advancers} 隊，請調整每組晉級名額或最佳名次補位數。"
        )


def build_advancement_data(
    groups: dict[str, list[str]],
    advance_per_group: int,
    wildcard_count: int,
    knockout_format: str = "semifinal",
) -> dict[str, Any]:
    total_advancers = (len(groups) * advance_per_group) + wildcard_count
    if wildcard_count == 0:
        summary = f"每組前 {advance_per_group} 名晉級，共 {total_advancers} 隊。"
    else:
        next_rank = advance_per_group + 1
        summary = (
            f"每組前 {advance_per_group} 名晉級，"
            f"另取最佳第 {next_rank} 名 {wildcard_count} 隊，共 {total_advancers} 隊。"
        )

    placeholders = []
    for group_name in groups:
        for rank in range(1, advance_per_group + 1):
            placeholders.append(f"{group_name}{rank}")
    for index in range(1, wildcard_count + 1):
        placeholders.append(f"Best{advance_per_group + 1}-{index}")

    return {
        "advance_per_group": advance_per_group,
        "wildcard_count": wildcard_count,
        "total_advancers": total_advancers,
        "knockout_format": knockout_format,
        "knockout_stage": knockout_format_label(knockout_format),
        "summary": summary,
        "placeholders": placeholders,
    }


def normalize_knockout_format(value: str) -> str:
    normalized = str(value).strip().lower()
    aliases = {
        "4": "semifinal",
        "four": "semifinal",
        "semi": "semifinal",
        "semis": "semifinal",
        "semifinal": "semifinal",
        "semifinals": "semifinal",
        "direct_semifinal": "semifinal",
        "8": "quarterfinal",
        "eight": "quarterfinal",
        "quarter": "quarterfinal",
        "quarterfinal": "quarterfinal",
        "quarterfinals": "quarterfinal",
    }
    if normalized not in aliases:
        raise ValueError("淘汰賽階段只能選擇直接四強或八強賽。")
    return aliases[normalized]


def knockout_advancer_count(knockout_format: str) -> int:
    return {"semifinal": 4, "quarterfinal": 8}[normalize_knockout_format(knockout_format)]


def knockout_format_label(knockout_format: str) -> str:
    return {"semifinal": "直接四強", "quarterfinal": "八強賽"}[normalize_knockout_format(knockout_format)]


def determine_schedule_mode(team_count: int, group_count: int, config: dict[str, Any]) -> str:
    if (
        team_count == int(config["template_schedule_team_count"])
        and group_count == int(config["template_schedule_group_count"])
    ):
        return "template_schedule"
    return "dynamic_schedule"


def normalize_download_options(download_options: dict[str, bool] | None = None) -> dict[str, bool]:
    normalized = dict(DEFAULT_DOWNLOAD_OPTIONS)
    if download_options is None:
        return normalized

    for key in AVAILABLE_OUTPUTS:
        if key in download_options:
            normalized[key] = bool(download_options[key])
    return normalized


def update_draw_runtime_options(
    draw_data: dict[str, Any],
    *,
    config: dict[str, Any] | None = None,
    download_options: dict[str, bool] | None = None,
    day1_latest_end: str | None = None,
    day2_latest_end: str | None = None,
) -> dict[str, Any]:
    config = config or load_config(BASE_DIR)
    updated = deepcopy(draw_data)
    if download_options is not None:
        updated["download_options"] = normalize_download_options(download_options)
    else:
        updated["download_options"] = normalize_download_options(updated.get("download_options"))

    if day1_latest_end is not None or day2_latest_end is not None or "schedule" not in updated:
        updated["schedule"] = build_schedule_data(
            updated,
            config,
            day1_latest_end=day1_latest_end,
            day2_latest_end=day2_latest_end,
        )
    return updated


def update_draw_referees(
    draw_data: dict[str, Any],
    referees: list[Any] | dict[str, Any] | str,
    *,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = config or load_config(BASE_DIR)
    updated = deepcopy(draw_data)
    updated["download_options"] = normalize_download_options(updated.get("download_options"))
    updated["download_options"]["json"] = True
    updated["download_options"]["schedule"] = True

    normalized_referees = normalize_referees(referees, updated.get("teams", []))
    assignments, warnings = assign_referees(updated, normalized_referees, config)
    updated["referees"] = normalized_referees
    updated["referee_assignments"] = assignments
    updated["referee_warnings"] = warnings
    updated["referees_per_match"] = int(config.get("referees_per_match", 3))
    return updated


def create_referee_only_artifacts(
    schedule_source: Path | str | BinaryIO,
    referees: list[Any] | dict[str, Any] | str,
    *,
    source_file: str = "uploaded_schedule.xlsx",
    config: dict[str, Any] | None = None,
    base_dir: Path = BASE_DIR,
) -> tuple[dict[str, Any], ArtifactPaths]:
    config = config or load_config(base_dir)
    template_bytes = None
    if not isinstance(schedule_source, (str, Path)) and hasattr(schedule_source, "read"):
        current_position = schedule_source.tell() if hasattr(schedule_source, "tell") else None
        template_bytes = schedule_source.read()
        if hasattr(schedule_source, "seek"):
            schedule_source.seek(current_position or 0)

    draw_data = load_existing_schedule_draw_data(schedule_source, source_file, config)
    template_copy = schedule_source if isinstance(schedule_source, (str, Path)) else None
    if template_copy is not None:
        draw_data["_existing_schedule_template_path"] = str(Path(template_copy).resolve())
    elif template_bytes is not None:
        draw_data["_existing_schedule_template_bytes"] = template_bytes
    draw_data = update_draw_referees(draw_data, referees, config=config)
    artifacts = generate_artifacts(draw_data, base_dir=base_dir)
    return draw_data, artifacts


def load_existing_schedule_draw_data(
    schedule_source: Path | str | BinaryIO,
    source_file: str,
    config: dict[str, Any],
) -> dict[str, Any]:
    workbook = load_workbook(schedule_source, data_only=False)
    matches = read_existing_schedule_matches(workbook, config)
    if not matches:
        raise ValueError("無法從這份 Excel 讀到賽程。請確認格式接近 113/114 賽程表，或包含工具產生的「賽程」工作表。")

    fields = list(config.get("fields", []))
    field_order = {field: index for index, field in enumerate(fields)}
    matches = sorted(matches, key=lambda match: match_sort_key(match, field_order))
    teams = sorted(
        {
            str(value).strip()
            for match in matches
            for value in (match.get("home"), match.get("away"))
            if is_real_team_name(value)
        }
    )

    schedule = {
        "status": "scheduled",
        "messages": [f"已從既有賽程讀取 {len(matches)} 場比賽，僅進行裁判排班。"],
        "warnings": [],
        "constraints": {
            "fields": fields,
            "match_duration_minutes": int(config.get("match_duration_minutes", 45)),
            "source": "existing_schedule",
        },
        "day_slots": summarize_match_time_slots(matches, fields),
        "match_count": len(matches),
        "matches": matches,
    }

    return {
        "drawn_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": source_file,
        "team_count": len(teams),
        "group_count": 0,
        "teams": teams,
        "slot_order": [],
        "slots": {},
        "groups": {},
        "advancement": {
            "advance_per_group": 0,
            "wildcard_count": 0,
            "total_advancers": 0,
            "summary": "既有賽程裁判排班模式，不重新抽籤或安排晉級。",
            "placeholders": [],
        },
        "knockout_format": "existing_schedule",
        "schedule_mode": "existing_schedule",
        "download_options": normalize_download_options({"json": True, "schedule": True, "pdf": False}),
        "random_function": "not_used_for_existing_schedule",
        "random_source": "既有賽程裁判排班模式未使用抽籤亂數。",
        "schedule": schedule,
    }


def read_existing_schedule_matches(workbook: Any, config: dict[str, Any]) -> list[dict[str, Any]]:
    for sheet in workbook.worksheets:
        dynamic_matches = read_dynamic_schedule_sheet(sheet, config)
        if dynamic_matches:
            return dynamic_matches

    for sheet in workbook.worksheets:
        template_matches = read_template_schedule_sheet(sheet, config)
        if template_matches:
            return template_matches

    return []


def read_dynamic_schedule_sheet(sheet: Any, config: dict[str, Any]) -> list[dict[str, Any]]:
    header_row = None
    for row_index in range(1, min(sheet.max_row, 25) + 1):
        values = [normalize_header_text(sheet.cell(row_index, column).value) for column in range(1, min(sheet.max_column, 12) + 1)]
        if any(value in {"match", "matchno", "matchnumber", "場次"} for value in values) and any(value in {"day", "日期", "天"} for value in values):
            header_row = row_index
            break

    if header_row is None:
        return []

    headers = {
        normalize_header_text(sheet.cell(header_row, column).value): column
        for column in range(1, min(sheet.max_column, 20) + 1)
        if sheet.cell(header_row, column).value is not None
    }

    def col(*names: str) -> int | None:
        for name in names:
            normalized = normalize_header_text(name)
            if normalized in headers:
                return headers[normalized]
        return None

    match_col = col("場次", "match", "matchno", "matchnumber")
    day_col = col("日期", "day")
    time_col = col("時間", "time")
    field_col = col("場地", "field")
    stage_col = col("階段", "stage")
    group_col = col("組別", "group")
    home_col = col("主隊", "隊伍1", "home", "home/label", "主隊/占位")
    away_col = col("客隊", "隊伍2", "away", "away/label", "客隊/占位")

    required = [match_col, day_col, time_col, field_col, home_col, away_col]
    if any(value is None for value in required):
        return []

    field_order = {field: index for index, field in enumerate(config.get("fields", []))}
    matches: list[dict[str, Any]] = []
    time_indexes: dict[tuple[str, str], int] = {}
    for row_index in range(header_row + 1, sheet.max_row + 1):
        match_no = parse_int_or_none(sheet.cell(row_index, match_col).value)
        if match_no is None:
            continue
        day = str(sheet.cell(row_index, day_col).value or "").strip()
        time_label = normalize_time_label(str(sheet.cell(row_index, time_col).value or "").strip())
        field = str(sheet.cell(row_index, field_col).value or "").strip()
        home = str(sheet.cell(row_index, home_col).value or "").strip()
        away = str(sheet.cell(row_index, away_col).value or "").strip()
        if not day or not time_label or not field or not home or not away:
            continue
        key = (day, time_label)
        if key not in time_indexes:
            time_indexes[key] = len({item[1] for item in time_indexes if item[0] == day})
        matches.append(
            build_existing_match(
                match_no=match_no,
                day=day,
                time=time_label,
                time_index=time_indexes[key],
                field=field,
                stage=str(sheet.cell(row_index, stage_col).value or "") if stage_col else "",
                group=str(sheet.cell(row_index, group_col).value or "") if group_col else "",
                home=home,
                away=away,
            )
        )

    return sorted(matches, key=lambda match: match_sort_key(match, field_order))


def read_template_schedule_sheet(sheet: Any, config: dict[str, Any]) -> list[dict[str, Any]]:
    field_order = {field: index for index, field in enumerate(config.get("fields", []))}
    time_indexes: dict[tuple[str, str], int] = {}
    matches: list[dict[str, Any]] = []
    for cell_map in TEMPLATE_SCHEDULE_MATCH_CELLS:
        match_no = parse_int_or_none(sheet[cell_map["match_cell"]].value)
        home = str(sheet[cell_map["home_cell"]].value or "").strip()
        away = str(sheet[cell_map["away_cell"]].value or "").strip()
        time_label = normalize_time_label(str(sheet[cell_map["time_cell"]].value or "").strip())
        if match_no is None or (not home and not away) or not time_label:
            continue
        day = cell_map["day"]
        key = (day, time_label)
        if key not in time_indexes:
            time_indexes[key] = len({item[1] for item in time_indexes if item[0] == day})
        group = str(sheet[cell_map["group_cell"]].value or "").strip() if cell_map["group_cell"] else ""
        matches.append(
            build_existing_match(
                match_no=match_no,
                day=day,
                time=time_label,
                time_index=time_indexes[key],
                field=cell_map["field"],
                stage=infer_stage_from_match_no(match_no),
                group=group,
                home=home,
                away=away,
            )
        )

    return sorted(matches, key=lambda match: match_sort_key(match, field_order))


def build_existing_match(
    *,
    match_no: int,
    day: str,
    time: str,
    time_index: int,
    field: str,
    stage: str,
    group: str,
    home: str,
    away: str,
) -> dict[str, Any]:
    return {
        "match_no": match_no,
        "day": day,
        "time": time,
        "time_index": time_index,
        "field": field,
        "stage": stage or infer_stage_from_match_no(match_no),
        "stage_code": infer_stage_code(stage, match_no),
        "group": group,
        "home": home,
        "away": away,
        "home_label": home,
        "away_label": away,
        "note": "",
    }


def summarize_match_time_slots(matches: list[dict[str, Any]], fields: list[str]) -> dict[str, list[dict[str, Any]]]:
    day_slots: dict[str, dict[tuple[int, str], dict[str, Any]]] = {}
    for match in matches:
        day = str(match.get("day", ""))
        key = (int(match.get("time_index", 0)), str(match.get("time", "")))
        day_slots.setdefault(day, {})
        day_slots[day].setdefault(
            key,
            {
                "day": day,
                "time_index": key[0],
                "time": key[1],
                "fields": [],
            },
        )
        field = match.get("field", "")
        if field and field not in day_slots[day][key]["fields"]:
            day_slots[day][key]["fields"].append(field)

    result: dict[str, list[dict[str, Any]]] = {}
    for day, slots in day_slots.items():
        result[day] = []
        for key in sorted(slots):
            item = slots[key]
            if fields:
                item["fields"] = sorted(item["fields"], key=lambda field: fields.index(field) if field in fields else 99)
            result[day].append(item)
    return result


def normalize_header_text(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "").replace("#", "")


def parse_int_or_none(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_time_label(value: str) -> str:
    return value.replace("–", "-").replace("—", "-").strip()


def infer_stage_from_match_no(match_no: int) -> str:
    if match_no <= 12:
        return "小組賽"
    if match_no <= 16:
        return "淘汰賽"
    if match_no <= 18:
        return "四強"
    return "決賽"


def infer_stage_code(stage: str, match_no: int) -> str:
    stage_text = str(stage or "")
    if "小組" in stage_text or match_no <= 12:
        return "group"
    if "季" in stage_text:
        return "third_place"
    if "冠" in stage_text or "決" in stage_text or match_no >= 19:
        return "final"
    if "四" in stage_text or match_no in {17, 18}:
        return "semifinal"
    return "knockout"


def is_real_team_name(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    marker_tokens = ("勝", "敗", "W", "L", "Best", "季殿", "冠亞")
    if any(token in text for token in marker_tokens):
        return False
    if len(text) <= 3 and text[0].isalpha() and any(ch.isdigit() for ch in text):
        return False
    return True


def normalize_referees(referees: list[Any] | dict[str, Any] | str, teams: list[str]) -> list[dict[str, Any]]:
    if isinstance(referees, dict):
        referees = referees.get("referees", [])
    elif isinstance(referees, str):
        referees = [{"name": line.strip()} for line in referees.splitlines() if line.strip()]

    if referees is None:
        referees = []

    valid_teams = set(teams)
    normalized: list[dict[str, Any]] = []
    seen_names: set[str] = set()

    for item in referees:
        if isinstance(item, str):
            raw_item = {"name": item}
        elif isinstance(item, dict):
            raw_item = dict(item)
        else:
            raise ValueError("裁判資料格式必須是姓名或包含 name 的物件。")

        name = str(raw_item.get("name", "")).strip()
        if not name:
            continue
        if name in seen_names:
            raise ValueError(f"裁判姓名重複：{name}")
        seen_names.add(name)

        affiliated_team = str(raw_item.get("affiliated_team") or raw_item.get("team") or "").strip()
        if affiliated_team and affiliated_team not in valid_teams:
            raise ValueError(f"{name} 的所屬隊伍不在本次隊伍名單內：{affiliated_team}")

        normalized.append(
            {
                "name": name,
                "affiliated_team": affiliated_team,
                "unavailable_match_nos": parse_match_numbers(raw_item.get("unavailable_match_nos", [])),
            }
        )

    return normalized


def parse_match_numbers(value: Any) -> list[int]:
    if value is None or value == "":
        return []

    if isinstance(value, str):
        normalized = value.replace("，", ",").replace("、", ",").replace(";", ",").replace("\n", ",")
        parts = [part.strip().lstrip("#") for part in normalized.split(",") if part.strip()]
    elif isinstance(value, (list, tuple, set)):
        parts = list(value)
    else:
        parts = [value]

    match_numbers: set[int] = set()
    for part in parts:
        try:
            match_no = int(part)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"不可排場次必須是數字：{part}") from exc
        if match_no <= 0:
            raise ValueError(f"不可排場次必須大於 0：{match_no}")
        match_numbers.add(match_no)
    return sorted(match_numbers)


def assign_referees(
    draw_data: dict[str, Any],
    referees: list[dict[str, Any]],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str]]:
    schedule = draw_data.get("schedule", {})
    matches = schedule.get("matches", [])
    referees_per_match = int(config.get("referees_per_match", 3))

    if schedule.get("status") != "scheduled" or not matches:
        return [], ["賽程尚未成功安排，無法產生裁判排班。"]
    if referees_per_match <= 0:
        return [], ["每場裁判人數必須大於 0。"]
    if not referees:
        return [], ["尚未輸入裁判名單，裁判表會保留空白。"]

    field_order = {field: index for index, field in enumerate(config.get("fields", []))}
    ordered_matches = sorted(matches, key=lambda match: match_sort_key(match, field_order))
    possible_teams_by_match = build_possible_teams_by_match(draw_data, ordered_matches)

    assignment_counts = {referee["name"]: 0 for referee in referees}
    day_counts: dict[tuple[str, str], int] = {}
    last_time_index: dict[tuple[str, str], int] = {}
    busy_by_time: dict[tuple[str, Any], set[str]] = {}
    assignments: list[dict[str, Any]] = []
    warnings: list[str] = []

    for match in ordered_matches:
        match_no = int(match.get("match_no", 0))
        day = str(match.get("day", ""))
        time_index = match.get("time_index", match.get("time", ""))
        busy_key = (day, time_index)
        busy_names = busy_by_time.setdefault(busy_key, set())
        possible_teams = possible_teams_by_match.get(match_no, set())
        assigned_names: list[str] = []

        for _slot_index in range(referees_per_match):
            candidate = choose_referee_candidate(
                referees,
                match_no=match_no,
                day=day,
                time_index=time_index,
                possible_teams=possible_teams,
                assigned_names=set(assigned_names),
                busy_names=busy_names,
                assignment_counts=assignment_counts,
                day_counts=day_counts,
                last_time_index=last_time_index,
            )
            if candidate is None:
                break

            name = candidate["name"]
            assigned_names.append(name)
            busy_names.add(name)
            assignment_counts[name] += 1
            day_counts[(name, day)] = day_counts.get((name, day), 0) + 1
            if isinstance(time_index, int):
                last_time_index[(name, day)] = time_index

        missing_count = referees_per_match - len(assigned_names)
        if missing_count > 0:
            warnings.append(f"第 {match_no} 場可用裁判不足，尚缺 {missing_count} 位。")

        assignments.append(
            {
                "match_no": match_no,
                "day": day,
                "time": match.get("time", ""),
                "time_index": time_index,
                "field": match.get("field", ""),
                "stage": match.get("stage", ""),
                "stage_code": match.get("stage_code", ""),
                "group": match.get("group", ""),
                "home": match.get("home", match.get("home_label", "")),
                "away": match.get("away", match.get("away_label", "")),
                "home_label": match.get("home_label", match.get("home", "")),
                "away_label": match.get("away_label", match.get("away", "")),
                "referees": assigned_names,
                "missing_count": missing_count,
                "possible_teams": sorted(possible_teams),
            }
        )

    return assignments, warnings


def choose_referee_candidate(
    referees: list[dict[str, Any]],
    *,
    match_no: int,
    day: str,
    time_index: Any,
    possible_teams: set[str],
    assigned_names: set[str],
    busy_names: set[str],
    assignment_counts: dict[str, int],
    day_counts: dict[tuple[str, str], int],
    last_time_index: dict[tuple[str, str], int],
) -> dict[str, Any] | None:
    best_referee: dict[str, Any] | None = None
    best_score: tuple[int, int, int, str] | None = None

    for referee in referees:
        name = referee["name"]
        if name in assigned_names or name in busy_names:
            continue
        if match_no in set(referee.get("unavailable_match_nos", [])):
            continue

        affiliated_team = referee.get("affiliated_team", "")
        if affiliated_team and affiliated_team in possible_teams:
            continue

        consecutive_penalty = 0
        if isinstance(time_index, int) and last_time_index.get((name, day)) == time_index - 1:
            consecutive_penalty = 1

        score = (
            assignment_counts.get(name, 0),
            day_counts.get((name, day), 0),
            consecutive_penalty,
            name,
        )
        if best_score is None or score < best_score:
            best_referee = referee
            best_score = score

    return best_referee


def match_sort_key(match: dict[str, Any], field_order: dict[str, int]) -> tuple[int, int, int, int]:
    day_order = {"DAY1": 1, "DAY2": 2}
    time_index = match.get("time_index")
    if not isinstance(time_index, int):
        time_index = 999
    return (
        day_order.get(str(match.get("day", "")), 99),
        time_index,
        field_order.get(match.get("field", ""), 99),
        int(match.get("match_no", 9999)),
    )


def build_possible_teams_by_match(
    draw_data: dict[str, Any],
    ordered_matches: list[dict[str, Any]],
) -> dict[int, set[str]]:
    groups = draw_data.get("groups", {})
    all_teams = set(draw_data.get("teams", []))
    token_possible: dict[str, set[str]] = {}
    possible_by_match: dict[int, set[str]] = {}

    for match in ordered_matches:
        match_no = int(match.get("match_no", 0))
        if match.get("stage_code") == "group":
            possible = {
                value
                for value in (match.get("home"), match.get("away"))
                if isinstance(value, str) and value
            }
        else:
            possible = set()
            for token in (match.get("home"), match.get("away"), match.get("home_label"), match.get("away_label")):
                possible.update(resolve_possible_teams(str(token or ""), groups, all_teams, token_possible))

        possible_by_match[match_no] = possible
        if match_no:
            token_possible[f"{match_no}W"] = set(possible)
            token_possible[f"{match_no}L"] = set(possible)
        for token_key in ("winner_label", "loser_label"):
            token = str(match.get(token_key) or "")
            if token:
                token_possible[token] = set(possible)

    return possible_by_match


def resolve_possible_teams(
    token: str,
    groups: dict[str, list[str]],
    all_teams: set[str],
    token_possible: dict[str, set[str]],
) -> set[str]:
    if not token:
        return set()
    if token in token_possible:
        return set(token_possible[token])
    if token.startswith("Best"):
        return set(all_teams)
    for group_name, group_teams in groups.items():
        if token.startswith(group_name) and token[len(group_name) :].isdigit():
            return set(group_teams)
    if token in all_teams:
        return {token}
    return set(all_teams)


def build_schedule_data(
    draw_data: dict[str, Any],
    config: dict[str, Any],
    *,
    day1_latest_end: str | None = None,
    day2_latest_end: str | None = None,
) -> dict[str, Any]:
    day1_latest_end = day1_latest_end or str(config["default_day1_latest_end"])
    day2_latest_end = day2_latest_end or str(config["default_day2_latest_end"])
    fields = list(config["fields"])
    duration = int(config["match_duration_minutes"])
    max_daily_matches = int(config["max_matches_per_team_per_day"])

    constraints = {
        "day1_latest_end": day1_latest_end,
        "day2_latest_end": day2_latest_end,
        "fields": fields,
        "match_duration_minutes": duration,
        "max_matches_per_team_per_day": max_daily_matches,
        "final_preferred_start": str(config["final_preferred_start"]),
    }

    day1_slots = build_field_slots("DAY1", day1_latest_end, config)
    day2_slots = build_field_slots("DAY2", day2_latest_end, config)
    messages: list[str] = []
    warnings: list[str] = []

    groups = draw_data["groups"]
    one_team_groups = [group_name for group_name, teams in groups.items() if len(teams) < 2]
    if one_team_groups:
        return infeasible_schedule(
            constraints,
            messages=[
                f"{'、'.join(one_team_groups)} 組只有 1 隊，無法安排完整小組賽。請減少組數或增加隊伍。",
            ],
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )

    oversized_groups = [group_name for group_name, teams in groups.items() if len(teams) - 1 > max_daily_matches]
    if oversized_groups:
        return infeasible_schedule(
            constraints,
            messages=[
                f"{'、'.join(oversized_groups)} 組隊伍數過多，完整單循環會讓同隊一天超過 {max_daily_matches} 場。",
            ],
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )

    group_matches = build_group_stage_matches(draw_data)
    if len(group_matches) > len(day1_slots):
        return infeasible_schedule(
            constraints,
            messages=[
                f"DAY1 可用場次為 {len(day1_slots)} 場，但完整小組賽需要 {len(group_matches)} 場。",
                "請在網頁把 DAY1 最晚結束時間延後，或調整組數。",
            ],
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )

    scheduled_group_matches, group_warning = schedule_group_matches(
        group_matches,
        day1_slots,
        max_daily_matches=max_daily_matches,
    )
    if len(scheduled_group_matches) != len(group_matches):
        return infeasible_schedule(
            constraints,
            messages=[
                "在目前 DAY1 時段與每日出賽上限下，無法排完完整小組賽。",
                "請延後 DAY1 最晚結束時間，或調整組數讓各組人數更平均。",
            ],
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )
    warnings.extend(group_warning)

    knockout_plan = build_knockout_matches(
        draw_data["advancement"]["placeholders"],
        max_daily_matches,
        draw_data.get("knockout_format", draw_data["advancement"].get("knockout_format", "semifinal")),
    )
    if knockout_plan["status"] != "scheduled":
        return infeasible_schedule(
            constraints,
            messages=knockout_plan["messages"],
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )

    scheduled_knockout, knockout_messages, knockout_warnings = schedule_knockout_matches(
        knockout_plan["rounds"],
        day2_slots,
        config,
    )
    if not scheduled_knockout and knockout_plan["match_count"] > 0:
        return infeasible_schedule(
            constraints,
            messages=knockout_messages,
            day1_slots=day1_slots,
            day2_slots=day2_slots,
        )
    warnings.extend(knockout_warnings)

    matches = assign_match_numbers(scheduled_group_matches + scheduled_knockout, fields=fields)
    messages.append(
        f"已排定小組賽 {len(scheduled_group_matches)} 場、淘汰賽 {len(scheduled_knockout)} 場。"
    )
    if draw_data.get("schedule_mode") == "template_schedule":
        messages.append("本次也符合 12 隊 4 組模板，Excel 仍會沿用 113 公開版格式。")

    return {
        "status": "scheduled",
        "messages": messages,
        "warnings": warnings,
        "constraints": constraints,
        "day_slots": {
            "DAY1": summarize_time_slots(day1_slots),
            "DAY2": summarize_time_slots(day2_slots),
        },
        "match_count": len(matches),
        "matches": matches,
    }


def infeasible_schedule(
    constraints: dict[str, Any],
    *,
    messages: list[str],
    day1_slots: list[dict[str, Any]],
    day2_slots: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "status": "infeasible",
        "messages": messages,
        "warnings": [
            "本工具不會為了塞進時段而刪減小組賽或製造不公平賽程。",
            "建議增加時段、調整組數、減少晉級隊數，或另外增加比賽日。",
        ],
        "constraints": constraints,
        "day_slots": {
            "DAY1": summarize_time_slots(day1_slots),
            "DAY2": summarize_time_slots(day2_slots),
        },
        "match_count": 0,
        "matches": [],
    }


def build_field_slots(day: str, latest_end: str, config: dict[str, Any]) -> list[dict[str, Any]]:
    duration = int(config["match_duration_minutes"])
    fields = list(config["fields"])
    starts = list(config["day_slots"][day])
    latest_end_minutes = time_to_minutes(latest_end)

    if not starts:
        return []

    next_start_minutes = time_to_minutes(starts[-1]) + 60
    while next_start_minutes + duration <= latest_end_minutes:
        starts.append(minutes_to_time(next_start_minutes))
        next_start_minutes += 60

    slots: list[dict[str, Any]] = []
    for time_index, start in enumerate(starts):
        time_label = build_time_label(start, duration)
        for field in fields:
            slots.append(
                {
                    "day": day,
                    "time_index": time_index,
                    "start": start,
                    "time": time_label,
                    "field": field,
                }
            )
    return slots


def summarize_time_slots(field_slots: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[tuple[str, str], dict[str, Any]] = {}
    for slot in field_slots:
        key = (slot["day"], slot["time"])
        if key not in seen:
            seen[key] = {
                "day": slot["day"],
                "time_index": slot["time_index"],
                "time": slot["time"],
                "fields": [],
            }
        seen[key]["fields"].append(slot["field"])
    return list(seen.values())


def build_group_stage_matches(draw_data: dict[str, Any]) -> list[dict[str, Any]]:
    team_to_slot = {team: slot for slot, team in draw_data["slots"].items()}
    matches: list[dict[str, Any]] = []
    for group_name, teams in draw_data["groups"].items():
        for round_index, home_index, away_index in round_robin_order(len(teams)):
            home = teams[home_index]
            away = teams[away_index]
            matches.append(
                {
                    "stage": "小組賽",
                    "stage_code": "group",
                    "group": group_name,
                    "round": round_index + 1,
                    "home": home,
                    "away": away,
                    "home_label": team_to_slot[home],
                    "away_label": team_to_slot[away],
                    "note": "",
                }
            )
    return sorted(matches, key=lambda item: (item["round"], item["group"]))


def round_robin_order(team_count: int) -> list[tuple[int, int, int]]:
    if team_count == 2:
        return [(0, 0, 1)]
    if team_count == 3:
        return [(0, 0, 1), (1, 1, 2), (2, 2, 0)]
    if team_count == 4:
        return [(0, 0, 3), (0, 1, 2), (1, 0, 2), (1, 3, 1), (2, 0, 1), (2, 2, 3)]

    return [
        (index, home_index, away_index)
        for index, (home_index, away_index) in enumerate(combinations(range(team_count), 2))
    ]


def schedule_group_matches(
    matches: list[dict[str, Any]],
    field_slots: list[dict[str, Any]],
    *,
    max_daily_matches: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    remaining = [deepcopy(match) for match in matches]
    scheduled: list[dict[str, Any]] = []
    warnings: list[str] = []
    team_day_counts: dict[tuple[str, str], int] = {}
    team_last_time_index: dict[tuple[str, str], int] = {}
    teams_by_day_time: dict[tuple[str, int], set[str]] = {}
    field_group_counts: dict[tuple[str, str], int] = {}

    for slot in field_slots:
        best_index: int | None = None
        best_score: tuple[int, int, int, str] | None = None
        occupied = teams_by_day_time.setdefault((slot["day"], slot["time_index"]), set())

        for index, match in enumerate(remaining):
            teams = {match["home"], match["away"]}
            if occupied.intersection(teams):
                continue
            if any(team_day_counts.get((team, slot["day"]), 0) >= max_daily_matches for team in teams):
                continue

            consecutive_penalty = sum(
                1
                for team in teams
                if team_last_time_index.get((team, slot["day"])) == slot["time_index"] - 1
            )
            field_penalty = field_group_counts.get((slot["field"], match["group"]), 0)
            load_penalty = sum(team_day_counts.get((team, slot["day"]), 0) for team in teams)
            score = (consecutive_penalty, field_penalty, load_penalty, f"{match['group']}-{match['round']}")
            if best_score is None or score < best_score:
                best_index = index
                best_score = score

        if best_index is None:
            continue

        match = remaining.pop(best_index)
        match.update(
            {
                "day": slot["day"],
                "time": slot["time"],
                "time_index": slot["time_index"],
                "field": slot["field"],
            }
        )
        scheduled.append(match)
        for team in (match["home"], match["away"]):
            team_day_counts[(team, slot["day"])] = team_day_counts.get((team, slot["day"]), 0) + 1
            team_last_time_index[(team, slot["day"])] = slot["time_index"]
            occupied.add(team)
        field_group_counts[(slot["field"], match["group"])] = field_group_counts.get((slot["field"], match["group"]), 0) + 1

        if best_score and best_score[0] > 0:
            warnings.append(
                f"{match['home']} vs {match['away']} 因時段有限，可能接近前一場出賽。"
            )

    return scheduled, warnings


def build_knockout_matches(placeholders: list[str], max_daily_matches: int, knockout_format: str = "semifinal") -> dict[str, Any]:
    participant_count = len(placeholders)
    knockout_format = normalize_knockout_format(knockout_format)
    if participant_count < 2:
        return {
            "status": "infeasible",
            "messages": ["晉級隊數少於 2 隊，無法安排淘汰賽。"],
            "rounds": [],
            "match_count": 0,
        }

    bracket_size = next_power_of_two(participant_count)
    if int(math.log2(bracket_size)) > max_daily_matches:
        return {
            "status": "infeasible",
            "messages": [
                f"本次晉級 {participant_count} 隊，淘汰賽路徑可能超過同隊一天 {max_daily_matches} 場上限。",
                "請減少晉級隊數到 8 隊以內，或另外增加淘汰賽比賽日。",
            ],
            "rounds": [],
            "match_count": 0,
        }

    rounds: list[dict[str, Any]] = []
    current = list(placeholders)
    match_index = 1

    if participant_count == 4:
        pairs = [(current[0], current[2]), (current[1], current[3])]
        round_matches, current, losers, match_index = make_knockout_round(
            pairs,
            stage="四強",
            stage_code="semifinal",
            match_index=match_index,
        )
        rounds.append({"stage": "四強", "stage_code": "semifinal", "matches": round_matches})
        semifinal_losers = losers
    elif participant_count == 8 and knockout_format == "quarterfinal":
        pairs = make_quarterfinal_pairs(current)
        round_matches, current, _losers, match_index = make_knockout_round(
            pairs,
            stage="八強",
            stage_code="knockout",
            match_index=match_index,
        )
        rounds.append({"stage": "八強", "stage_code": "knockout", "matches": round_matches})

        pairs = [(current[0], current[1]), (current[2], current[3])]
        round_matches, current, losers, match_index = make_knockout_round(
            pairs,
            stage="四強",
            stage_code="semifinal",
            match_index=match_index,
        )
        rounds.append({"stage": "四強", "stage_code": "semifinal", "matches": round_matches})
        semifinal_losers = losers
    else:
        semifinal_losers: list[str] = []
        first_round_match_count = participant_count - (bracket_size // 2)
        if first_round_match_count > 0:
            bye_count = participant_count - (first_round_match_count * 2)
            bye_participants = current[:bye_count]
            play_in_participants = current[bye_count:]
            pairs = make_seed_pairs(play_in_participants)
            stage = stage_name_for_count(len(play_in_participants))
            round_matches, winners, _losers, match_index = make_knockout_round(
                pairs,
                stage=stage,
                stage_code="knockout",
                match_index=match_index,
            )
            rounds.append({"stage": stage, "stage_code": "knockout", "matches": round_matches})
            current = bye_participants + winners

        while len(current) > 2:
            pairs = make_seed_pairs(current)
            stage = stage_name_for_count(len(current))
            stage_code = "semifinal" if len(current) == 4 else "knockout"
            round_matches, current, losers, match_index = make_knockout_round(
                pairs,
                stage=stage,
                stage_code=stage_code,
                match_index=match_index,
            )
            rounds.append({"stage": stage, "stage_code": stage_code, "matches": round_matches})
            if stage_code == "semifinal":
                semifinal_losers = losers

    final_matches = [
        {
            "stage": "季軍賽",
            "stage_code": "third_place",
            "home": semifinal_losers[0] if len(semifinal_losers) > 0 else "四強敗者1",
            "away": semifinal_losers[1] if len(semifinal_losers) > 1 else "四強敗者2",
            "home_label": semifinal_losers[0] if len(semifinal_losers) > 0 else "四強敗者1",
            "away_label": semifinal_losers[1] if len(semifinal_losers) > 1 else "四強敗者2",
            "note": "",
        },
        {
            "stage": "冠軍賽",
            "stage_code": "final",
            "home": current[0],
            "away": current[1],
            "home_label": current[0],
            "away_label": current[1],
            "note": "",
        },
    ]
    if participant_count == 2:
        final_matches = [final_matches[1]]

    rounds.append({"stage": "決賽", "stage_code": "finals", "matches": final_matches})
    return {
        "status": "scheduled",
        "messages": [],
        "rounds": rounds,
        "match_count": sum(len(round_item["matches"]) for round_item in rounds),
    }


def make_seed_pairs(participants: list[str]) -> list[tuple[str, str]]:
    return [
        (participants[index], participants[-(index + 1)])
        for index in range(len(participants) // 2)
    ]


def make_quarterfinal_pairs(participants: list[str]) -> list[tuple[str, str]]:
    participant_set = set(participants)
    preferred_pairs = [("A1", "B2"), ("B1", "A2"), ("C1", "D2"), ("D1", "C2")]
    if len(participants) == 8 and all(item in participant_set for pair in preferred_pairs for item in pair):
        return preferred_pairs

    return [
        (participants[index], participants[index + 1])
        for index in range(0, len(participants), 2)
    ]


def make_knockout_round(
    pairs: list[tuple[str, str]],
    *,
    stage: str,
    stage_code: str,
    match_index: int,
) -> tuple[list[dict[str, Any]], list[str], list[str], int]:
    matches = []
    winners = []
    losers = []
    for home, away in pairs:
        winner = f"K{match_index}W"
        loser = f"K{match_index}L"
        matches.append(
            {
                "stage": stage,
                "stage_code": stage_code,
                "home": home,
                "away": away,
                "home_label": home,
                "away_label": away,
                "winner_label": winner,
                "loser_label": loser,
                "note": "",
            }
        )
        winners.append(winner)
        losers.append(loser)
        match_index += 1
    return matches, winners, losers, match_index


def stage_name_for_count(participant_count: int) -> str:
    if participant_count == 2:
        return "冠軍賽"
    if participant_count == 4:
        return "四強"
    if participant_count == 8:
        return "八強"
    if participant_count == 16:
        return "十六強"
    return "淘汰賽第一輪"


def schedule_knockout_matches(
    rounds: list[dict[str, Any]],
    field_slots: list[dict[str, Any]],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    if not rounds:
        return [], ["沒有可排的淘汰賽場次。"], []

    fields = list(config["fields"])
    final_preferred_start = str(config["final_preferred_start"])
    time_groups = group_slots_by_time(field_slots)
    scheduled: list[dict[str, Any]] = []
    warnings: list[str] = []
    next_time_index = 0
    rounds_before_finals = [round_item for round_item in rounds if round_item["stage_code"] != "finals"]
    final_round = next((round_item for round_item in rounds if round_item["stage_code"] == "finals"), None)

    for round_item in rounds_before_finals:
        round_matches = [deepcopy(match) for match in round_item["matches"]]
        chunk_size = len(fields)
        chunks = [round_matches[index : index + chunk_size] for index in range(0, len(round_matches), chunk_size)]
        for chunk in chunks:
            if next_time_index >= len(time_groups):
                return [], [
                    f"DAY2 可用時段不足，無法排完 {round_item['stage']}。",
                    "請在網頁把 DAY2 最晚結束時間延後，或減少晉級隊數。",
                ], warnings
            time_slot = time_groups[next_time_index]
            for match, field in zip(chunk, fields):
                match.update(
                    {
                        "day": "DAY2",
                        "time": time_slot["time"],
                        "time_index": time_slot["time_index"],
                        "field": field,
                        "group": "",
                    }
                )
                scheduled.append(match)
            next_time_index += 1

    if final_round is None:
        return scheduled, [], warnings

    final_matches = [deepcopy(match) for match in final_round["matches"]]
    preferred_index = find_time_index(time_groups, final_preferred_start)
    earliest_final_index = max(preferred_index, next_time_index)
    rest_friendly_index = earliest_final_index + 1 if rounds_before_finals and earliest_final_index == next_time_index else earliest_final_index
    final_time_index = find_available_final_index(time_groups, rest_friendly_index, len(final_matches), fields)
    if final_time_index is None:
        final_time_index = find_available_final_index(time_groups, earliest_final_index, len(final_matches), fields)
    if final_time_index is None:
        return [], [
            "DAY2 沒有足夠場地安排季軍賽與冠軍賽。",
            "請延後 DAY2 最晚結束時間，或減少晉級隊數。",
        ], warnings

    if final_time_index > preferred_index:
        warnings.append("冠軍賽/季軍賽因前面淘汰賽需要休息或時段不足，已自動後移。")

    final_time_slot = time_groups[final_time_index]
    for match, field in zip(final_matches, fields):
        match.update(
            {
                "day": "DAY2",
                "time": final_time_slot["time"],
                "time_index": final_time_slot["time_index"],
                "field": field,
                "group": "",
            }
        )
        scheduled.append(match)

    return scheduled, [], warnings


def group_slots_by_time(field_slots: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[int, dict[str, Any]] = {}
    for slot in field_slots:
        grouped.setdefault(
            slot["time_index"],
            {"time_index": slot["time_index"], "time": slot["time"], "fields": []},
        )
        grouped[slot["time_index"]]["fields"].append(slot["field"])
    return [grouped[index] for index in sorted(grouped)]


def find_time_index(time_groups: list[dict[str, Any]], start_time: str) -> int:
    for index, item in enumerate(time_groups):
        if item["time"].startswith(start_time):
            return index
    return 0


def find_available_final_index(
    time_groups: list[dict[str, Any]],
    start_index: int,
    match_count: int,
    fields: list[str],
) -> int | None:
    for index in range(start_index, len(time_groups)):
        if len(fields) >= match_count:
            return index
    return None


def assign_match_numbers(matches: list[dict[str, Any]], fields: list[str] | None = None) -> list[dict[str, Any]]:
    field_order = {field: index for index, field in enumerate(fields or [])}
    sorted_matches = sorted(
        matches,
        key=lambda item: (
            item.get("day", ""),
            int(item.get("time_index", 0)),
            field_order.get(item.get("field", ""), 99),
            str(item.get("field", "")),
            stage_sort_order(str(item.get("stage_code", ""))),
        ),
    )
    for index, match in enumerate(sorted_matches, start=1):
        match["match_no"] = index
    return sorted_matches


def stage_sort_order(stage_code: str) -> int:
    return {
        "group": 0,
        "knockout": 1,
        "semifinal": 2,
        "third_place": 3,
        "final": 4,
    }.get(stage_code, 9)


def next_power_of_two(value: int) -> int:
    if value <= 1:
        return 1
    return 1 << (value - 1).bit_length()


def time_to_minutes(value: str) -> int:
    hour_text, minute_text = value.split(":")
    return (int(hour_text) * 60) + int(minute_text)


def minutes_to_time(minutes: int) -> str:
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def build_time_label(start: str, duration: int) -> str:
    return f"{start}-{minutes_to_time(time_to_minutes(start) + duration)}"


def load_draw_data(draw_path: Path) -> dict[str, Any]:
    with draw_path.open("r", encoding="utf-8") as file_handle:
        draw_data = json.load(file_handle)
    validate_draw_data(draw_data)
    return draw_data


def validate_draw_data(draw_data: dict[str, Any]) -> None:
    required_keys = {
        "drawn_at",
        "source_file",
        "teams",
        "slots",
        "groups",
        "random_function",
    }
    missing_keys = required_keys - set(draw_data)
    if missing_keys:
        missing_list = ", ".join(sorted(missing_keys))
        raise ValueError(f"抽籤結果缺少必要欄位：{missing_list}")

    if draw_data.get("schedule_mode") == "existing_schedule":
        if not draw_data.get("schedule", {}).get("matches"):
            raise ValueError("既有賽程模式缺少比賽資料。")
        return

    groups = draw_data["groups"]
    if not isinstance(groups, dict) or not groups:
        raise ValueError("抽籤結果沒有分組資料。")

    flattened = [team for group_teams in groups.values() for team in group_teams]
    if len(flattened) != len(draw_data["teams"]):
        raise ValueError("抽籤結果隊伍總數與分組內容不一致。")
    if len(set(flattened)) != len(flattened):
        raise ValueError("抽籤結果有重複隊伍，請重新確認 draw_result.json。")


def create_draw_artifacts(
    base_dir: Path = BASE_DIR,
    *,
    registration_source: Path | str | BinaryIO | None = None,
    source_file: str | None = None,
    group_count: int | None = None,
    advance_per_group: int | None = None,
    wildcard_count: int | None = None,
    knockout_format: str | None = None,
    download_options: dict[str, bool] | None = None,
    day1_latest_end: str | None = None,
    day2_latest_end: str | None = None,
) -> tuple[dict[str, Any], ArtifactPaths]:
    config = load_config(base_dir)

    if registration_source is None:
        registration_path = resolve_registration_path(base_dir)
        registration_source = registration_path
        source_file = registration_path.name

    teams = load_teams(registration_source, config=config)
    draw_data = create_draw_data(
        teams,
        source_file or "uploaded.xlsx",
        group_count=int(config["default_group_count"]) if group_count is None else group_count,
        advance_per_group=int(config["default_advance_per_group"]) if advance_per_group is None else advance_per_group,
        wildcard_count=int(config["default_wildcard_count"]) if wildcard_count is None else wildcard_count,
        knockout_format=knockout_format,
        config=config,
        download_options=download_options,
        day1_latest_end=day1_latest_end,
        day2_latest_end=day2_latest_end,
    )
    artifacts = generate_artifacts(draw_data, base_dir=base_dir)
    return draw_data, artifacts


def clear_latest_artifacts(base_dir: Path = BASE_DIR) -> bool:
    latest_dir = (base_dir / "outputs" / "latest").resolve()
    if not latest_dir.exists():
        return True

    blocked = False
    for child_path in latest_dir.iterdir():
        try:
            if child_path.is_dir():
                shutil.rmtree(child_path)
            else:
                child_path.unlink()
        except PermissionError:
            blocked = True

    if not blocked:
        try:
            latest_dir.rmdir()
        except OSError:
            blocked = True

    return not blocked


def generate_artifacts(
    draw_data: dict[str, Any],
    *,
    base_dir: Path = BASE_DIR,
    output_dir: Path | None = None,
    outputs: dict[str, bool] | None = None,
) -> ArtifactPaths:
    validate_draw_data(draw_data)

    draw_data = deepcopy(draw_data)
    draw_data["download_options"] = normalize_download_options(outputs or draw_data.get("download_options"))

    config = load_config(base_dir)
    artifact_filenames = get_artifact_filenames(base_dir)
    timestamp_dir_name = make_timestamp_dirname(draw_data["drawn_at"])

    if output_dir is None:
        output_dir = ARCHIVE_ROOT / timestamp_dir_name
    else:
        output_dir = output_dir.resolve()

    latest_dir = LATEST_ROOT.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    latest_dir.mkdir(parents=True, exist_ok=True)

    draw_json_path = output_dir / DRAW_RESULT_FILENAME
    schedule_path = output_dir / artifact_filenames["schedule"] if draw_data["download_options"]["schedule"] else None
    pdf_path = output_dir / artifact_filenames["pdf"] if draw_data["download_options"]["pdf"] else None

    latest_draw_json_path = latest_dir / DRAW_RESULT_FILENAME
    latest_schedule_path = latest_dir / artifact_filenames["schedule"] if draw_data["download_options"]["schedule"] else None
    latest_pdf_path = latest_dir / artifact_filenames["pdf"] if draw_data["download_options"]["pdf"] else None

    write_draw_result(clean_draw_data_for_json(draw_data), draw_json_path)
    json_synced = sync_latest_copy(draw_json_path, latest_draw_json_path)
    schedule_synced = True
    pdf_synced = True

    if schedule_path is not None and latest_schedule_path is not None:
        build_schedule_output(draw_data, schedule_path, config, base_dir)
        schedule_synced = sync_latest_copy(schedule_path, latest_schedule_path)
    else:
        remove_latest_if_exists(latest_dir / artifact_filenames["schedule"])

    if pdf_path is not None and latest_pdf_path is not None:
        build_randomness_pdf(draw_data, pdf_path, config)
        pdf_synced = sync_latest_copy(pdf_path, latest_pdf_path)
    else:
        remove_latest_if_exists(latest_dir / artifact_filenames["pdf"])

    return ArtifactPaths(
        output_dir=output_dir,
        latest_dir=latest_dir,
        draw_json_path=draw_json_path,
        schedule_path=schedule_path,
        pdf_path=pdf_path,
        latest_draw_json_path=latest_draw_json_path,
        latest_schedule_path=latest_schedule_path,
        latest_pdf_path=latest_pdf_path,
        latest_sync_complete=json_synced and schedule_synced and pdf_synced,
    )


def make_timestamp_dirname(timestamp_text: str) -> str:
    return timestamp_text.replace(":", "").replace("-", "").replace("T", "_")


def write_draw_result(draw_data: dict[str, Any], draw_json_path: Path) -> None:
    with draw_json_path.open("w", encoding="utf-8") as file_handle:
        json.dump(draw_data, file_handle, ensure_ascii=False, indent=2)


def clean_draw_data_for_json(draw_data: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in draw_data.items() if not key.startswith("_")}


def sync_latest_copy(source_path: Path, target_path: Path) -> bool:
    if source_path.resolve() == target_path.resolve():
        return True
    try:
        shutil.copy2(source_path, target_path)
    except PermissionError:
        return False
    return True


def remove_latest_if_exists(target_path: Path) -> bool:
    try:
        if target_path.exists():
            target_path.unlink()
    except PermissionError:
        return False
    return True


def generate_pdf_only(
    draw_data: dict[str, Any],
    *,
    base_dir: Path = BASE_DIR,
    output_dir: Path | None = None,
) -> tuple[Path, Path, bool]:
    validate_draw_data(draw_data)

    config = load_config(base_dir)
    artifact_filenames = get_artifact_filenames(base_dir)
    timestamp_dir_name = make_timestamp_dirname(draw_data["drawn_at"])

    if output_dir is None:
        output_dir = ARCHIVE_ROOT / timestamp_dir_name
    else:
        output_dir = output_dir.resolve()

    latest_dir = LATEST_ROOT.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    latest_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = output_dir / artifact_filenames["pdf"]
    latest_pdf_path = latest_dir / artifact_filenames["pdf"]

    build_randomness_pdf(draw_data, pdf_path, config)
    pdf_synced = sync_latest_copy(pdf_path, latest_pdf_path)
    return pdf_path, latest_pdf_path, pdf_synced


def build_schedule_output(
    draw_data: dict[str, Any],
    output_path: Path,
    config: dict[str, Any],
    base_dir: Path,
) -> None:
    if draw_data.get("schedule_mode") == "existing_schedule":
        build_existing_schedule_workbook(draw_data, output_path, config)
        return

    if draw_data.get("schedule_mode") == "template_schedule":
        template_path = resolve_template_path(base_dir)
        build_template_schedule_workbook(draw_data, template_path, output_path, config)
        return

    build_dynamic_schedule_workbook(draw_data, output_path, config)


def build_existing_schedule_workbook(draw_data: dict[str, Any], output_path: Path, config: dict[str, Any]) -> None:
    template_path_text = draw_data.get("_existing_schedule_template_path")
    if template_path_text and Path(template_path_text).exists():
        workbook = load_workbook(Path(template_path_text))
    elif draw_data.get("_existing_schedule_template_bytes"):
        workbook = load_workbook(io.BytesIO(draw_data["_existing_schedule_template_bytes"]))
    else:
        workbook = Workbook()
        schedule_sheet = workbook.active
        schedule_sheet.title = "賽程"
        write_schedule_sheet(schedule_sheet, draw_data)

    write_referee_sheet_if_needed(workbook, draw_data, config)
    workbook.save(output_path)


def build_template_schedule_workbook(
    draw_data: dict[str, Any],
    template_path: Path,
    output_path: Path,
    config: dict[str, Any],
) -> None:
    workbook = load_workbook(template_path)
    sheet = workbook.worksheets[0]
    sheet["A1"] = config["schedule_title"]

    groups = draw_data["groups"]
    for group_name in TEMPLATE_GROUPS:
        summary_cells = GROUP_SUMMARY_CELLS[group_name]
        teams = groups[group_name]
        sheet[summary_cells[0]] = str(config["group_label_format"]).format(group=group_name)
        for cell_name, team_name in zip(summary_cells[1:], teams):
            sheet[cell_name] = team_name

        match_cells = GROUP_MATCH_CELLS[group_name]
        match_order = GROUP_MATCH_ORDERS[group_name]
        for (home_cell, away_cell), (home_index, away_index) in zip(match_cells, match_order):
            sheet[home_cell] = teams[home_index]
            sheet[away_cell] = teams[away_index]

    if normalize_knockout_format(draw_data.get("knockout_format", "semifinal")) == "quarterfinal":
        apply_cell_values(sheet, build_template_quarterfinal_labels(draw_data))
    else:
        apply_cell_values(sheet, config["semifinal_labels"])
        for cell_name in config["clear_cells"]:
            sheet[cell_name] = None

        apply_cell_values(sheet, config["final_labels"])
        for cell_name in config["reserve_cells"]:
            sheet[cell_name] = None

    write_referee_sheet_if_needed(workbook, draw_data, config)
    workbook.save(output_path)


def build_template_quarterfinal_labels(draw_data: dict[str, Any]) -> dict[str, Any]:
    placeholders = list(draw_data["advancement"]["placeholders"])
    if len(placeholders) != 8:
        raise ValueError("八強模板需要剛好 8 隊晉級。")

    quarterfinal_pairs = make_quarterfinal_pairs(placeholders)
    return {
        "B14": 13,
        "D14": quarterfinal_pairs[0][0],
        "E14": quarterfinal_pairs[0][1],
        "G14": 14,
        "I14": quarterfinal_pairs[1][0],
        "J14": quarterfinal_pairs[1][1],
        "B15": 15,
        "D15": quarterfinal_pairs[2][0],
        "E15": quarterfinal_pairs[2][1],
        "G15": 16,
        "I15": quarterfinal_pairs[3][0],
        "J15": quarterfinal_pairs[3][1],
        "B17": 17,
        "D17": "13W",
        "E17": "14W",
        "G17": 18,
        "I17": "15W",
        "J17": "16W",
        "B18": 19,
        "D18": "17L",
        "E18": "18L",
        "G18": 20,
        "I18": "17W",
        "J18": "18W",
    }


def build_dynamic_schedule_workbook(draw_data: dict[str, Any], output_path: Path, config: dict[str, Any]) -> None:
    workbook = Workbook()
    group_sheet = workbook.active
    group_sheet.title = "分組結果"
    schedule_sheet = workbook.create_sheet("賽程")
    warning_sheet = workbook.create_sheet("排程說明")

    write_group_sheet(group_sheet, draw_data, config)
    write_schedule_sheet(schedule_sheet, draw_data)
    write_warning_sheet(warning_sheet, draw_data)
    write_referee_sheet_if_needed(workbook, draw_data, config)
    workbook.save(output_path)


def write_group_sheet(sheet: Any, draw_data: dict[str, Any], config: dict[str, Any]) -> None:
    title = str(config["schedule_title"]).replace("賽程表", "分組結果")
    sheet["A1"] = title
    sheet["A2"] = f"抽籤時間：{draw_data['drawn_at']}"
    sheet["A3"] = f"報名來源：{draw_data['source_file']}"
    sheet["A4"] = f"晉級規則：{draw_data['advancement']['summary']}"
    sheet["A1"].font = Font(bold=True, size=16)

    header_fill = PatternFill("solid", fgColor="DDEBDD")
    header_font = Font(bold=True)
    thin_border = Border(bottom=Side(style="thin", color="A8B9A8"))

    groups = draw_data["groups"]
    max_group_size = max(len(teams) for teams in groups.values())
    start_row = 6

    for column_index, (group_name, teams) in enumerate(groups.items(), start=1):
        column_letter = get_column_letter(column_index)
        header_cell = sheet.cell(row=start_row, column=column_index)
        header_cell.value = str(config["group_label_format"]).format(group=group_name)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.border = thin_border
        header_cell.alignment = Alignment(horizontal="center")

        for row_offset in range(max_group_size):
            team_cell = sheet.cell(row=start_row + 1 + row_offset, column=column_index)
            team_cell.value = teams[row_offset] if row_offset < len(teams) else ""
            team_cell.alignment = Alignment(horizontal="center")

        sheet.column_dimensions[column_letter].width = 18

    advancement_row = start_row + max_group_size + 3
    sheet.cell(row=advancement_row, column=1).value = "晉級占位"
    sheet.cell(row=advancement_row, column=1).font = header_font
    for index, placeholder in enumerate(draw_data["advancement"]["placeholders"], start=1):
        sheet.cell(row=advancement_row + index, column=1).value = placeholder


def write_schedule_sheet(sheet: Any, draw_data: dict[str, Any]) -> None:
    schedule = draw_data.get("schedule", {})
    sheet["A1"] = "自動賽程"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A2"] = f"排程狀態：{schedule.get('status', 'unknown')}"

    headers = ["場次", "日期", "時間", "場地", "階段", "組別", "主隊/占位", "客隊/占位", "備註"]
    start_row = 4
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBDD")
        cell.alignment = Alignment(horizontal="center")

    matches = schedule.get("matches", [])
    if not matches:
        sheet.cell(row=start_row + 1, column=1).value = "目前沒有可輸出的賽程。請查看「排程說明」。"
    else:
        for row_index, match in enumerate(matches, start=start_row + 1):
            values = [
                match.get("match_no", ""),
                match.get("day", ""),
                match.get("time", ""),
                match.get("field", ""),
                match.get("stage", ""),
                match.get("group", ""),
                match.get("home", match.get("home_label", "")),
                match.get("away", match.get("away_label", "")),
                match.get("note", ""),
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index).value = value

    widths = [8, 10, 16, 8, 12, 8, 22, 22, 28]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def write_referee_sheet_if_needed(workbook: Any, draw_data: dict[str, Any], config: dict[str, Any]) -> None:
    if "referees" not in draw_data and "referee_assignments" not in draw_data:
        return

    sheet_name = str(config.get("referee_sheet_name", REFEREE_SHEET_NAME))
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(sheet_name)
    write_referee_sheet(sheet, draw_data, config)


def write_referee_sheet(sheet: Any, draw_data: dict[str, Any], config: dict[str, Any]) -> None:
    title = str(config.get("schedule_title", "")).replace("賽程表", "裁判表")
    sheet["A1"] = title or "裁判表"
    sheet["A1"].font = Font(bold=True, size=16)

    assignments = {int(item.get("match_no", 0)): item for item in draw_data.get("referee_assignments", [])}
    matches = draw_data.get("schedule", {}).get("matches", [])
    fields = list(config.get("fields", []))[:2]
    if len(fields) < 2:
        fields = fields + [f"Field {index + 1}" for index in range(len(fields), 2)]

    header_fill = PatternFill("solid", fgColor="DDEBDD")
    section_fill = PatternFill("solid", fgColor="EFF6EF")
    thin_border = Border(bottom=Side(style="thin", color="A8B9A8"))
    center = Alignment(horizontal="center", vertical="center")

    current_row = 2
    for day in ordered_schedule_days(draw_data):
        time_rows = referee_time_rows_for_day(draw_data, day)
        if not time_rows:
            continue

        sheet.cell(row=current_row, column=2).value = day
        sheet.cell(row=current_row, column=2).font = Font(bold=True, size=13)
        current_row += 1

        for column_index, value in ((3, fields[0]), (7, fields[1])):
            cell = sheet.cell(row=current_row, column=column_index)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.fill = section_fill
        current_row += 1

        for time_row in time_rows:
            sheet.cell(row=current_row, column=1).value = time_row["time"]
            sheet.cell(row=current_row, column=1).alignment = center

            for field_index, field in enumerate(fields):
                start_column = 2 if field_index == 0 else 6
                match = find_match_for_referee_row(matches, day, time_row, field)
                if match is None:
                    continue

                match_no = int(match.get("match_no", 0))
                assignment = assignments.get(match_no, {})
                sheet.cell(row=current_row, column=start_column).value = match_no
                sheet.cell(row=current_row, column=start_column).alignment = center
                for offset in range(int(config.get("referees_per_match", 3))):
                    referee_name = ""
                    if offset < len(assignment.get("referees", [])):
                        referee_name = assignment["referees"][offset]
                    sheet.cell(row=current_row, column=start_column + 1 + offset).value = referee_name
                    sheet.cell(row=current_row, column=start_column + 1 + offset).alignment = center

            current_row += 1

        current_row += 1

    warning_row = current_row + 1
    referee_warnings = list(draw_data.get("referee_warnings", []))
    if referee_warnings:
        sheet.cell(row=warning_row, column=1).value = "裁判排班提醒"
        sheet.cell(row=warning_row, column=1).font = Font(bold=True)
        for offset, warning in enumerate(referee_warnings, start=1):
            sheet.cell(row=warning_row + offset, column=1).value = warning
        current_row = warning_row + len(referee_warnings) + 2

    if draw_data.get("referees"):
        counts = referee_assignment_counts(draw_data.get("referee_assignments", []))
        sheet.cell(row=current_row, column=1).value = "裁判分配統計"
        sheet.cell(row=current_row, column=1).font = Font(bold=True)
        sheet.cell(row=current_row + 1, column=1).value = "姓名"
        sheet.cell(row=current_row + 1, column=2).value = "所屬隊伍"
        sheet.cell(row=current_row + 1, column=3).value = "分配場數"
        for column_index in range(1, 4):
            cell = sheet.cell(row=current_row + 1, column=column_index)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
        for row_offset, referee in enumerate(draw_data["referees"], start=2):
            sheet.cell(row=current_row + row_offset, column=1).value = referee["name"]
            sheet.cell(row=current_row + row_offset, column=2).value = referee.get("affiliated_team", "")
            sheet.cell(row=current_row + row_offset, column=3).value = counts.get(referee["name"], 0)

    widths = [16, 8, 16, 16, 16, 8, 16, 16, 16, 18, 18]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    for row in sheet.iter_rows(min_row=1, max_row=max(sheet.max_row, 1), min_col=1, max_col=9):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border


def ordered_schedule_days(draw_data: dict[str, Any]) -> list[str]:
    day_slots = draw_data.get("schedule", {}).get("day_slots", {})
    days = list(day_slots)
    for match in draw_data.get("schedule", {}).get("matches", []):
        day = str(match.get("day", ""))
        if day and day not in days:
            days.append(day)
    return sorted(days, key=lambda day: {"DAY1": 1, "DAY2": 2}.get(day, 99))


def referee_time_rows_for_day(draw_data: dict[str, Any], day: str) -> list[dict[str, Any]]:
    day_slots = draw_data.get("schedule", {}).get("day_slots", {})
    rows = list(day_slots.get(day, []))
    if rows:
        return rows

    grouped: dict[tuple[Any, str], dict[str, Any]] = {}
    for match in draw_data.get("schedule", {}).get("matches", []):
        if match.get("day") != day:
            continue
        key = (match.get("time_index", 999), str(match.get("time", "")))
        grouped.setdefault(
            key,
            {
                "day": day,
                "time_index": match.get("time_index", 999),
                "time": match.get("time", ""),
                "fields": [],
            },
        )
        if match.get("field") not in grouped[key]["fields"]:
            grouped[key]["fields"].append(match.get("field"))
    return [grouped[key] for key in sorted(grouped)]


def find_match_for_referee_row(
    matches: list[dict[str, Any]],
    day: str,
    time_row: dict[str, Any],
    field: str,
) -> dict[str, Any] | None:
    for match in matches:
        if match.get("day") != day:
            continue
        if match.get("field") != field:
            continue
        if "time_index" in time_row and match.get("time_index") == time_row.get("time_index"):
            return match
        if match.get("time") == time_row.get("time"):
            return match
    return None


def referee_assignment_counts(assignments: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for assignment in assignments:
        for referee_name in assignment.get("referees", []):
            counts[referee_name] = counts.get(referee_name, 0) + 1
    return counts


def write_warning_sheet(sheet: Any, draw_data: dict[str, Any]) -> None:
    schedule = draw_data.get("schedule", {})
    sheet["A1"] = "排程說明"
    sheet["A1"].font = Font(bold=True, size=16)
    lines = [
        f"狀態：{schedule.get('status', 'unknown')}",
        "",
        "訊息：",
        *schedule.get("messages", []),
        "",
        "提醒：",
        *schedule.get("warnings", []),
    ]
    for row_index, line in enumerate(lines, start=2):
        sheet.cell(row=row_index, column=1).value = line
    sheet.column_dimensions["A"].width = 90


def apply_cell_values(sheet: Any, cell_values: dict[str, Any]) -> None:
    for cell_name, value in cell_values.items():
        sheet[cell_name] = value


def build_randomness_pdf(
    draw_data: dict[str, Any],
    output_path: Path,
    config: dict[str, Any] | None = None,
) -> None:
    config = config or load_config(BASE_DIR)
    pdf_title = str(config["pdf_title"])
    font_name = register_pdf_font()
    pdf_canvas = canvas.Canvas(str(output_path), pagesize=A4)
    margin_x = 48
    current_y = A4[1] - 54

    pdf_canvas.setTitle(pdf_title)
    pdf_canvas.setFont(font_name, 18)
    pdf_canvas.drawString(margin_x, current_y, pdf_title)
    current_y -= 30

    schedule = draw_data.get("schedule", {})
    intro_lines = [
        "一、抽籤目的",
        f"本次共有 {draw_data['team_count']} 隊參與抽籤，程式只讀取報名表隊名，並以完全隨機方式分組。",
        "",
        "二、使用的隨機函數",
        f"本工具使用 Python 函數：{RANDOM_FUNCTION_NAME}",
        "SystemRandom 會使用作業系統提供的安全亂數來源，不使用固定種子，也不做人為排序。",
        f"亂數來源摘要：{RANDOM_FUNCTION_SUMMARY}。",
        "",
        "三、分組流程",
        build_slot_assignment_summary(draw_data),
        "洗牌完成後，隊伍會依序輪流放入各組，因此各組隊伍數會盡量平均。",
        "",
        "四、賽制與排程摘要",
        f"本次分成 {draw_data['group_count']} 組。",
        f"晉級規則：{draw_data['advancement']['summary']}",
        f"淘汰賽階段：{draw_data['advancement'].get('knockout_stage', knockout_format_label(draw_data.get('knockout_format', 'semifinal')))}",
        build_template_knockout_summary_line(draw_data, config) if draw_data.get("schedule_mode") == "template_schedule" else "非 12 隊 4 組時，工具會另產生動態賽程 Excel。",
        f"排程狀態：{schedule.get('status', 'unknown')}",
        *schedule.get("messages", []),
        "",
        "五、本次抽籤紀錄",
        f"抽籤時間：{draw_data['drawn_at']}",
        f"報名來源：{draw_data['source_file']}",
    ]

    current_y = draw_wrapped_lines(
        pdf_canvas,
        intro_lines,
        font_name=font_name,
        start_x=margin_x,
        start_y=current_y,
        font_size=11,
        max_chars=44,
        line_height=17,
    )

    current_y -= 8
    group_lines = ["六、分組結果"]
    for group_name, teams in draw_data["groups"].items():
        team_text = "、".join(teams)
        group_lines.append(f"{group_name}組：{team_text}")

    draw_wrapped_lines(
        pdf_canvas,
        group_lines,
        font_name=font_name,
        start_x=margin_x,
        start_y=current_y,
        font_size=11,
        max_chars=44,
        line_height=17,
    )

    pdf_canvas.save()


def build_slot_assignment_summary(draw_data: dict[str, Any]) -> str:
    first_slots = "、".join(draw_data["slot_order"][: min(12, len(draw_data["slot_order"]))])
    if len(draw_data["slot_order"]) > 12:
        first_slots += "..."
    return f"洗牌後依序填入的籤位為：{first_slots}。"


def build_semifinal_summary_line(config: dict[str, Any]) -> str:
    labels = config["semifinal_labels"]
    left_pair = f"{format_stage_label(labels['D14'])} 對 {format_stage_label(labels['E14'])}"
    right_pair = f"{format_stage_label(labels['I14'])} 對 {format_stage_label(labels['J14'])}"
    return f"四強固定為：{left_pair}、{right_pair}。"


def build_template_knockout_summary_line(draw_data: dict[str, Any], config: dict[str, Any]) -> str:
    if normalize_knockout_format(draw_data.get("knockout_format", "semifinal")) == "quarterfinal":
        return "八強賽沿用 113 公開版模板：10:00 與 11:00 打八強，14:00 打四強，15:00 打季軍與冠軍。"
    return build_semifinal_summary_line(config)


def format_stage_label(label: Any) -> str:
    text = str(label).strip()
    if len(text) == 2 and text[0] in GROUP_LABELS and text[1].isdigit():
        return f"{text[0]}組第{text[1]}"
    return text


def draw_wrapped_lines(
    pdf_canvas: canvas.Canvas,
    lines: list[str],
    *,
    font_name: str,
    start_x: int,
    start_y: float,
    font_size: int,
    max_chars: int,
    line_height: int,
) -> float:
    text_object = pdf_canvas.beginText(start_x, start_y)
    text_object.setFont(font_name, font_size)

    for line in lines:
        if not line:
            text_object.textLine("")
            continue
        for wrapped_line in wrap_cjk_text(line, max_chars):
            text_object.textLine(wrapped_line)

    pdf_canvas.drawText(text_object)
    line_count = sum(max(1, len(wrap_cjk_text(line, max_chars))) for line in lines)
    return start_y - (line_count * line_height)


def wrap_cjk_text(text: str, max_chars: int) -> list[str]:
    chunks: list[str] = []
    remaining = text.strip()
    while remaining:
        chunks.append(remaining[:max_chars])
        remaining = remaining[max_chars:]
    return chunks or [""]


def register_pdf_font() -> str:
    if PDF_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return PDF_FONT_NAME

    for font_path in PDF_FONT_CANDIDATES:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, str(font_path)))
                return PDF_FONT_NAME
            except Exception:
                continue

    fallback_font_name = "STSong-Light"
    if fallback_font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont(fallback_font_name))
    return fallback_font_name


def get_latest_draw_data(base_dir: Path = BASE_DIR) -> dict[str, Any] | None:
    latest_path = base_dir / "outputs" / "latest" / DRAW_RESULT_FILENAME
    if not latest_path.exists():
        return None
    return load_draw_data(latest_path)
